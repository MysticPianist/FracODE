### This program fits a dataset to a system of three coupled fractional ODEs using a gradient descent algorithm
## Euler's method is used to calculate both the solutions to the ODEs and to get the partial derivatives at each time value
# Julian Frank, Dr. Mark Allen 8/20/2023

import time
from openpyxl import Workbook, load_workbook # Only needed if extracting data from excel (copy this code into another Python IDE)
import numpy as np
import matplotlib.pyplot as plt
from scipy.special import gamma
from scipy.integrate import quad

# define system of Caputo ODEs. u = [S, I, R]
def caputo_odes(t, u, params):

    # more constants
    a = params[0]
    b = params[1]
    c = params[2]

    # assign each ODE to a vector element
    S = u[0]
    I = u[1]
    R = u[2]

    # system of ODEs
    dSdt = -a*S*I + c*R
    dIdt = a*S*I - b*I
    dRdt = b*I - c*R

    return [dSdt, dIdt, dRdt]

# function to solve the frac ode system. Takes in ODEs, alpha value, desired solution domain, and array of initial values
# returns multidimentional array of solution values corresponding with t
def solve_fractional_ode_system(caputo_odes, params, t, u):
    # define step size
    h = t[1] - t[0]
    a = params[3]

    # iterate for every time t, except for the initial value
    for i in range(1, len(t)):
        # Get the n value based on the already calvulated u values
        n = i - 1

         # Summation portion
        z = np.zeros(n)
        for k in range(n):
            z[k] = (n+1-k)**(1-a) - (n-k)**(1-a)

        s_sum = np.sum([(u[j+1, 0] - u[j, 0])*(z[j]) for j in range(n)])
        i_sum = np.sum([(u[j+1, 1] - u[j, 1])*(z[j]) for j in range(n)])
        r_sum = np.sum([(u[j+1, 2] - u[j, 2])*(z[j]) for j in range(n)])

        # gamma portion from pdf
        gamma_part = (1-a)* gamma(1-a) * h**a
        caputo_ode = caputo_odes(i, u[n], params[:3])
        s_gamma_part = gamma_part * caputo_ode[0]
        i_gamma_part = gamma_part * caputo_ode[1]
        r_gamma_part = gamma_part * caputo_ode[2]

        # calculate u(t_{n+1})
        u[i, 0] = (u[i-1, 0] - s_sum + s_gamma_part)
        u[i, 1] = (u[i-1, 1] - i_sum + i_gamma_part)
        u[i, 2] = (u[i-1, 2] - r_sum + r_gamma_part)

    return u

def get_partials(x, t, params):
    h = t[1] - t[0]
    # Define partial derivatives
    # S
    dSda = np.zeros(len(t))
    dSdb = np.zeros(len(t))
    dSdc = np.zeros(len(t))
    dSdu = np.zeros(len(t))
    # I
    dIda = np.zeros(len(t))
    dIdb = np.zeros(len(t))
    dIdc = np.zeros(len(t))
    dIdu = np.zeros(len(t))
    # R
    dRda = np.zeros(len(t))
    dRdb = np.zeros(len(t))
    dRdc = np.zeros(len(t))
    dRdu = np.zeros(len(t))

    S = x[:, 0]
    I = x[:, 1]
    R = x[:, 2]

    a = params[0]
    b = params[1]
    c = params[2]
    u = params[3]

    ### calculate partial derivatives numerically
    gamma_prime = quad(lambda t: (t**(-u)) * (np.exp(-t)) * (np.log(t)), 0, 20)[0]
    gamma_prime = round(gamma_prime, 10)
    
    for i in range(len(t)-1):
        n = i - 1

        z = np.zeros(n+1)
        w = np.zeros(n+1)
        for k in range(n):
            # storing common expressions to be indexed for efficiency
            z[k] = (n+1-k)**(1-u) - (n-k)**(1-u)
            w[k] = (((n-k)**(1-u))*(np.log(n-k))) - ((n+1-k)**(1-u)) * (np.log(n+1-k))
        
        universal_gamma_part = ((gamma(1-u))*(-h**(u) + (1-u)*(h**(u))*np.log(u)) - ((1-u)*(h**(u)) * gamma_prime))
        universal_gamma_part_2 = ((1-u) * gamma(1-u) * (h**(u)))

        # partials of S
        dSdu_sum_part = dSdu[i] - np.sum([((dSdu[j+1] - dSdu[j])*z[j] + ((S[j+1] - S[j])*(w[j]))) for j in range(n)])
        dSdu_gamma_part = (-a*S[i]*I[i] + c*R[i]) * universal_gamma_part + universal_gamma_part_2 * (-a*((dSdu[i] * I[i]) + (S[i] * dIdu[i])) + c*dRdu[i])
        
        dSdu[i+1] = dSdu_sum_part + dSdu_gamma_part
        dSda[i+1] = dSda[i] - np.sum([(dSda[j+1] - dSda[j])*(z[j]) for j in range(n)]) + universal_gamma_part_2 * (-(S[i]*I[i] + a*(S[i]*dIda[i] + I[i]*dSda[i]))+c*dRda[i]) 
        dSdb[i+1] = dSdb[i] - np.sum([(dSdb[j+1] - dSdb[j])*(z[j]) for j in range(n)]) + universal_gamma_part_2 * (-a*(S[i]*dIdb[i] + I[i]*dSdb[i]) + c*dRdb[i])
        dSdc[i+1] = dSdc[i] - np.sum([(dSdc[j+1] - dSdc[j])*(z[j]) for j in range(n)]) + universal_gamma_part_2 * (-a*(S[i]*dIdc[i] + I[i]*dSdc[i]) + R[i] + c*dRdc[i])
        
        # partials of I
        dIdu_sum_part = dIdu[i] - np.sum([((dIdu[j+1] - dIdu[j])*(z[j]) + ((I[j+1] - I[j])*(w[j]))) for j in range(n)])
        dIdu_gamma_part = (a*S[i]*I[i] - b*I[i]) * universal_gamma_part + universal_gamma_part_2 * (a*((dSdu[i] * I[i]) + (S[i] * dIdu[i])) + b*dIdu[i])
        
        dIdu[i+1] = dIdu_sum_part + dIdu_gamma_part
        dIda[i+1] = dIda[i] - np.sum([(dIda[j+1] - dIda[j])*(z[j]) for j in range(n)]) + universal_gamma_part_2 * (S[i]*I[i] + a*((S[i])*dIda[i]+I[i]*dSda[i]) - b*dIda[i]) 
        dIdb[i+1] = dIdb[i] - np.sum([(dIdb[j+1] - dIdb[j])*(z[j]) for j in range(n)]) + universal_gamma_part_2 * (a*(S[i]*dIdb[i] + I[i]*dSdb[i]) - (I[i] + b*dIdb[i]))
        dIdc[i+1] = dIdc[i] - np.sum([(dIdc[j+1] - dIdc[j])*(z[j]) for j in range(n)]) + universal_gamma_part_2 * (a*(S[i]*dIdc[i] + I[i]*dSdc[i]) - b*dIdc[i])
        
        # partials of R
        dRdu_sum_part = dRdu[i] - np.sum([((dRdu[j+1] - dRdu[j])*(z[j]) + ((R[j+1] - R[j])*(w[j]))) for j in range(n)])
        dRdu_gamma_part = (b*I[i] - c*R[i]) * universal_gamma_part + universal_gamma_part_2 * (b*dIdu[i] - c*dRdu[i])
        
        dRdu[i+1] = dRdu_sum_part + dRdu_gamma_part
        dRda[i+1] = dRda[i] - np.sum([(dRda[j+1] - dRda[j])*(z[j]) for j in range(n)]) + universal_gamma_part_2 * (b*dIda[i] - c*dRda[i])
        dRdb[i+1] = dRdb[i] - np.sum([(dRdb[j+1] - dRdb[j])*(z[j]) for j in range(n)]) + universal_gamma_part_2 * ((I[i] + b*dIdb[i]) - c*dRdb[i])
        dRdc[i+1] = dRdc[i] - np.sum([(dRdc[j+1] - dRdc[j])*(z[j]) for j in range(n)]) + universal_gamma_part_2 * (b*dIdc[i] - (R[i] + c*dRdc[i]))
    
    partials = np.array([dIda, dIdb, dIdc, dIdu])

    return partials

def grad_desc(t, x, xhat, params, partials, learning_rate, time_step):
    # Unpack partials
    dIda = partials[0]
    dIdb = partials[1]
    dIdc = partials[2]
    dIdu = partials[3]

    # Define loss functions
    dlda_I = 0.0
    dldb_I = 0.0
    dldc_I = 0.0
    dldu_I = 0.0

    # calculate the losses
    n = len(x)
    for i in range(1, n+1):
        # Calculate gradient of the loss function with respect to a
        dlda_I += 2*((xhat[(i-1)*time_step, 1]-x[i-1]))*dIda[((i-1)*time_step)]
        # Calculate gradient of the loss function with respect to b
        dldb_I += 2*((xhat[(i-1)*time_step, 1]-x[i-1]))*dIdb[((i-1)*time_step)]
        # Calculate gradient of the loss function with respect to c
        dldc_I += 2*((xhat[(i-1)*time_step, 1]-x[i-1]))*dIdc[((i-1)*time_step)]
        # Calculate gradient of the loss function with respect to alpha
        dldu_I += 2*((xhat[(i-1)*time_step, 1]-x[i-1]))*dIdu[((i-1)*time_step)]

    ## Sum loss function
    loss_sum_I = np.sum([((xhat[(time_step*z), 1]-x[z])**2) for z in range(n)])

    # Update the parameters in the opposite direction of the partial derivative of the loss function
    params[0] = params[0] - learning_rate*(1/n)*dlda_I
    params[1] = params[1] - learning_rate*(1/n)*dldb_I
    params[2] = params[2] - learning_rate*(1/n)*dldc_I

    # alpha direction (adjust magnifier porportionally to how quickly it should change)
    params[3] = params[3] - (1000)*learning_rate*(1/n)*dldu_I

    # calculate the loss average
    loss_I = loss_sum_I/n

    return np.array([params[0], params[1], params[2], params[3], dldu_I, loss_I])

def get_random_params(min, max):
    a = np.random.uniform(min, max)
    b = np.random.uniform(min, max)
    c = np.random.uniform(min, max)
    u = np.random.uniform(.2, 1.0)
    
    return np.array([a, b, c, u])

start_time = time.time()  # Record start time

### This doesn't work in Jupyter Notebooks so only use it if copying this into another IDE
# get data from excel sheet
# wb = load_workbook('test_data.xlsx') # change as needed
# ws = wb.active
# x = []
# for row in range(2, 12):
    # x.append(ws['A' + str(row)].value)

#######################################                                              ##############################################
#######################################             DATA TO FIT GOES BELOW           ##############################################
#######################################                                              ##############################################
# manually input data
# x = [99, 95, 91, 85, 80, 74, 69, 69, 69, 69, 69, 69, 62, 54, 38, 25, 12, 6, 3, 1]
# x = [1, 3, 6, 12, 25, 38, 54, 62, 69, 72, 73, 74, 72, 69, 62, 54, 38, 30, 30, 30]
x = [1, 2, 4, 7, 13, 17, 23, 31, 41, 56, 64, 69, 73, 77, 82, 85, 88, 91, 92, 93]
# x = [99, 78, 54]
# x = [1, 15, 39, 83, 125, 142, 164, 182, 205, 220]
# x = [1, 15, 39, 83, 125, 142, 164, 150, 130, 100]
t = np.linspace(0, 19, len(x))

# timestep for euler's method (10 is fairly accurate for populations below 200) Warning: Expect longer execution time if set higher
time_step = 5

# Initial guess for model
t_hat = np.linspace(0, 19, (len(x)-1)*time_step+1)

def monte_carlo_method(params_guess):
    xhat = np.zeros((len(t_hat), 3))
    xhat[0] = np.array([99, 1, 0])
    learning_rate = 0.0000000005
    for epoch in range(3):
        xhat = solve_fractional_ode_system(caputo_odes, params_guess, t_hat, xhat)
        xhat_partials = get_partials(xhat, t_hat, params_guess)
        descent = grad_desc(t, x, xhat, params_guess, xhat_partials, learning_rate, time_step)
        params_guess = descent[:4]
        loss = descent[5]
    
    return (loss, params_guess)

# params_guess = np.array([0.009768952749205119, 0.044670475065208035, 0.023377978899198557, 0.7497920638081791]) # solved to fit default data
# define model population and initial values
xhat = np.zeros((len(t_hat), 3))
xhat[0] = np.array([99, 1, 0])

# scales the epoch optimization rate (adjust as needed)
learning_rate = 0.0000000005

## Optimization process
# using Monte-Carlo method to find initial parameter guess
parameter_options = {}
for i in range(100):
    params_guess = monte_carlo_method(get_random_params(.01, .05))
    if str(params_guess[0]) != 'nan':
        parameter_options[params_guess[0]] = params_guess[1]
    
min_loss = min(parameter_options.keys())
    
params_guess = parameter_options[min_loss]
# print(parameter_options)
# print("Minimum:", params_guess, min_loss)
# print(parameter_options.keys())

# main recursion
for epoch in range(1001):
    xhat = solve_fractional_ode_system(caputo_odes, params_guess, t_hat, xhat)
    xhat_partials = get_partials(xhat, t_hat, params_guess)
    descent = grad_desc(t, x, xhat, params_guess, xhat_partials, learning_rate, time_step)
    params_guess = descent[:5]
    loss = descent[5]
    # print(f'{epoch} dldu_I is {params_guess[4]}')
    print(f'{epoch} I loss is {loss}')
    # print(f'parameter b: {params_guess[1]}')
    # print(f'parameter c: {params_guess[2]}')
    # print(f'parameter alpha: {params_guess[3]}')
    if loss < 1:
        learning_rate = 0.0000000003
    elif loss < .5:
        learning_rate = 0.0000000002
        
end_time = time.time()    # Record end time
execution_time = end_time - start_time
print(f"Execution time: {execution_time} seconds")

# print and plot results
print(f'{epoch} dldu_I is {params_guess[4]}')
print(f'{epoch} I loss is {loss}\nparameter a: {params_guess[0]}')
print(f'parameter b: {params_guess[1]}')
print(f'parameter c: {params_guess[2]}')
print(f'parameter alpha: {params_guess[3]}')
plt.plot(t_hat, xhat[:, 0], label='S_hat')
plt.plot(t_hat, xhat[:, 1], label='I_hat')
plt.plot(t_hat, xhat[:, 2], label='R_hat')
plt.plot(t, x, label='I')
plt.legend()
plt.show()